# -*- coding: utf-8 -*-
"""
Excel导出服务
负责将检测记录和统计数据导出为Excel文件
"""

import os
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import settings


class ExcelExportService:
    """Excel导出服务"""
    
    def __init__(self):
        self.temp_dir = settings.export_temp_dir if hasattr(settings, 'export_temp_dir') else "./temp/exports"
        os.makedirs(self.temp_dir, exist_ok=True)
    
    async def export_detection_records(
        self,
        records: List[Dict],
        filepath: Optional[str] = None,
        user_name: str = "用户",
        filename_prefix: str = "检测记录"
    ) -> str:
        """
        导出检测记录为Excel文件
        
        Args:
            records: 检测记录列表
            filepath: 指定的文件路径（可选）
            user_name: 用户名
            filename_prefix: 文件名前缀
            
        Returns:
            str: Excel文件路径
        """
        if not records:
            raise ValueError("没有可导出的记录")
        
        # 创建工作簿
        wb = self._create_detection_workbook(records)
        
        # 确定文件路径
        if filepath:
            # 使用指定路径,规范化路径分隔符
            filepath = os.path.normpath(filepath)
        else:
            # 生成默认文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{user_name}_{timestamp}.xlsx"
            filepath = os.path.normpath(os.path.join(self.temp_dir, filename))
        
        # 确保目录存在
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # 保存文件
        wb.save(filepath)
        
        return filepath
    
    def _create_detection_workbook(self, records: List[Dict]) -> Workbook:
        """创建检测记录工作簿"""
        wb = Workbook()
        ws = wb.active
        ws.title = "检测记录"
        
        # 定义表头
        headers = ["序号", "检测时间", "图片名称", "病害类型", "置信度", "防治建议"]
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._style_header_cell(cell)
        
        # 写入数据
        for row_num, record in enumerate(records, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=self._format_datetime(record.get('created_at')))
            ws.cell(row=row_num, column=3, value=record.get('image_name', ''))
            ws.cell(row=row_num, column=4, value=record.get('disease_type', ''))
            ws.cell(row=row_num, column=5, value=self._format_confidence(record.get('confidence')))
            ws.cell(row=row_num, column=6, value=record.get('treatment_advice', ''))
            
            # 设置数据行样式
            for col_num in range(1, 7):
                self._style_data_cell(ws.cell(row=row_num, column=col_num))
        
        # 调整列宽
        self._adjust_column_widths(ws, headers)
        
        return wb
    
    def _style_header_cell(self, cell):
        """设置表头单元格样式"""
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _style_data_cell(self, cell):
        """设置数据单元格样式"""
        cell.font = Font(name='微软雅黑', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
    
    def _adjust_column_widths(self, ws, headers):
        """调整列宽"""
        column_widths = {
            "序号": 8,
            "检测时间": 20,
            "图片名称": 25,
            "病害类型": 15,
            "置信度": 12,
            "防治建议": 50
        }
        
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = column_widths.get(header, 15)
    
    def _format_datetime(self, dt) -> str:
        """格式化日期时间"""
        if isinstance(dt, datetime):
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(dt, str):
            return dt
        return ""
    
    def _format_confidence(self, confidence) -> str:
        """格式化置信度"""
        if confidence is None:
            return ""
        try:
            return f"{float(confidence):.2%}"
        except (ValueError, TypeError):
            return str(confidence)
    
    async def export_statistics(
        self,
        statistics: Dict,
        filename_prefix: str = "统计报表"
    ) -> str:
        """
        导出统计报表为Excel文件
        
        Args:
            statistics: 统计数据字典
            filename_prefix: 文件名前缀
            
        Returns:
            str: Excel文件路径
        """
        wb = self._create_statistics_workbook(statistics)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        filepath = os.path.join(self.temp_dir, filename)
        
        # 保存文件
        wb.save(filepath)
        
        return filepath
    
    def _create_statistics_workbook(self, statistics: Dict) -> Workbook:
        """创建统计报表工作簿"""
        wb = Workbook()
        
        # 病害类型统计
        if 'disease_types' in statistics:
            ws = wb.active
            ws.title = "病害类型统计"
            self._write_disease_type_stats(ws, statistics['disease_types'])
        
        # 时间趋势统计
        if 'time_trends' in statistics:
            ws = wb.create_sheet("时间趋势")
            self._write_time_trend_stats(ws, statistics['time_trends'])
        
        # 用户活跃度统计
        if 'user_activity' in statistics:
            ws = wb.create_sheet("用户活跃度")
            self._write_user_activity_stats(ws, statistics['user_activity'])
        
        return wb
    
    def _write_disease_type_stats(self, ws, data: List[Dict]):
        """写入病害类型统计"""
        headers = ["病害类型", "检测次数", "占比"]
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._style_header_cell(cell)
        
        # 写入数据
        for row_num, item in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=item.get('disease_type', ''))
            ws.cell(row=row_num, column=2, value=item.get('count', 0))
            ws.cell(row=row_num, column=3, value=self._format_confidence(item.get('percentage')))
            
            for col_num in range(1, 4):
                self._style_data_cell(ws.cell(row=row_num, column=col_num))
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _write_time_trend_stats(self, ws, data: List[Dict]):
        """写入时间趋势统计"""
        headers = ["日期", "检测次数"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._style_header_cell(cell)
        
        for row_num, item in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=item.get('date', ''))
            ws.cell(row=row_num, column=2, value=item.get('count', 0))
            
            for col_num in range(1, 3):
                self._style_data_cell(ws.cell(row=row_num, column=col_num))
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
    
    def _write_user_activity_stats(self, ws, data: List[Dict]):
        """写入用户活跃度统计"""
        headers = ["用户名", "检测次数", "最后检测时间"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._style_header_cell(cell)
        
        for row_num, item in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=item.get('username', ''))
            ws.cell(row=row_num, column=2, value=item.get('count', 0))
            ws.cell(row=row_num, column=3, value=self._format_datetime(item.get('last_detection')))
            
            for col_num in range(1, 4):
                self._style_data_cell(ws.cell(row=row_num, column=col_num))
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20


# 单例实例
_export_service_instance = None


def get_export_service() -> ExcelExportService:
    """获取导出服务单例"""
    global _export_service_instance
    if _export_service_instance is None:
        _export_service_instance = ExcelExportService()
    return _export_service_instance
